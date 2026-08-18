"""Build the registered 5 main and 3 supplementary manuscript tables."""

from __future__ import annotations

import os
import tempfile
import textwrap
from pathlib import Path

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image

from final_common import (
    MAIN_TABLE_DIR,
    PROJECT_ROOT,
    QC_DIR,
    SUPP_TABLE_DIR,
    TABLE_SOURCE_ROOT,
    ensure_directories,
    font_properties,
    mm_to_in,
    publication_style,
    relative_to_final,
    sha256_file,
    write_json,
    write_text,
)


TABLE_DEFINITIONS = [
    {
        "id": "Table1",
        "scope": "main",
        "source": TABLE_SOURCE_ROOT / "main_tables" / "table1_variables.csv",
        "stem": "Table01_clustering_variable_definitions",
        "title": "Clustering-variable definitions and transformations",
        "note": (
            "Variables were defined for the formal 790-event sample. MLLCL was "
            "transformed with log(1+x) before standardization; signed SRH values "
            "were retained."
        ),
    },
    {
        "id": "Table2",
        "scope": "main",
        "source": TABLE_SOURCE_ROOT / "main_tables" / "table2_k_metrics.csv",
        "stem": "Table02_candidate_cluster_number_diagnostics",
        "title": "Candidate cluster-number diagnostics",
        "note": (
            "The diagnostics describe solutions with k=2–6. No single diagnostic "
            "is interpreted as determining a uniquely preferred partition."
        ),
    },
    {
        "id": "Table3",
        "scope": "main",
        "source": TABLE_SOURCE_ROOT / "main_tables" / "table3_regime_characteristics.csv",
        "stem": "Table03_k3_regime_characteristics",
        "title": "Median characteristics of the three k=3 environmental regimes",
        "note": (
            "Values are medians in original units. Regime counts are C0=131, "
            "C1=307, and C2=352."
        ),
    },
    {
        "id": "Table4",
        "scope": "main",
        "source": TABLE_SOURCE_ROOT / "main_tables" / "table4_stability.csv",
        "stem": "Table04_stability_algorithm_sensitivity",
        "title": "Stability and algorithm-sensitivity metrics",
        "note": (
            "ARI denotes adjusted Rand index. Event stability categories describe "
            "assignment consistency; boundary events are not treated as errors."
        ),
    },
    {
        "id": "Table5",
        "scope": "main",
        "source": TABLE_SOURCE_ROOT / "main_tables" / "table5_weather_association.csv",
        "stem": "Table05_weather_type_association",
        "title": "Post-hoc weather-type association statistics",
        "note": (
            "Statistics use the 3×9 table with 787 valid weather-type records. "
            "The interval is the registered 95% bootstrap confidence interval."
        ),
    },
    {
        "id": "TableS1",
        "scope": "supplementary",
        "source": TABLE_SOURCE_ROOT / "supplementary_tables" / "tableS1_k4_statistics.csv",
        "stem": "TableS01_k4_cluster_statistics",
        "title": "k=4 cluster statistics",
        "note": (
            "Cluster labels remain numerical. Values are medians in original units "
            "for the accepted k=4 partition."
        ),
    },
    {
        "id": "TableS2",
        "scope": "supplementary",
        "source": TABLE_SOURCE_ROOT / "supplementary_tables" / "tableS2_sensitivity_complete.csv",
        "stem": "TableS02_complete_structural_sensitivity_results",
        "title": "Complete structural-sensitivity results",
        "note": (
            "ARI comparisons are relative to the corresponding accepted k-means "
            "solution. These analyses describe structural sensitivity and algorithm dependence."
        ),
    },
    {
        "id": "TableS3",
        "scope": "supplementary",
        "source": TABLE_SOURCE_ROOT / "supplementary_tables" / "tableS3_stpmod.csv",
        "stem": "TableS03_stpmod_rated_event_results",
        "title": "Exploratory STP_mod rated-event results",
        "note": (
            "The comparison includes only rated events; unrated events are excluded. "
            "Results are research-only and do not define a decision rule or probability."
        ),
    },
]


def _load_and_format(definition: dict) -> tuple[pd.DataFrame, str]:
    source = definition["source"]
    if not source.is_file():
        raise FileNotFoundError(source)
    data = pd.read_csv(source)
    table_id = definition["id"]
    transformation_note = "Directly formatted from the accepted registered table."

    if table_id == "Table1":
        output = data.rename(
            columns={
                "Variable": "Variable",
                "Unit": "Unit",
                "Transform": "Clustering transform",
                "Meaning": "Definition",
            }
        ).fillna("None")
        output["Unit"] = output["Unit"].replace(
            {"J/kg": "J kg−1", "m/s": "m s−1", "m²/s²": "m² s−2"}
        )
        output["Clustering transform"] = output["Clustering transform"].replace(
            {"None (signed)": "None; signed values retained"}
        )
    elif table_id == "Table2":
        output = data.copy()
        output.columns = [
            "k",
            "Silhouette",
            "Davies–Bouldin",
            "Calinski–Harabasz",
            "Minimum cluster n",
            "Maximum cluster n",
        ]
        for column in ("Silhouette", "Davies–Bouldin", "Calinski–Harabasz"):
            output[column] = output[column].map(lambda value: f"{value:.3f}")
    elif table_id == "Table3":
        output = data.copy()
        output.columns = [
            "Regime",
            "n",
            "2-m dew point (K)",
            "MLCAPE (J kg−1)",
            "MLLCL (m)",
            "0–6-km shear (m s−1)",
            "0–1-km SRH (m² s−2)",
        ]
        for column in output.columns[2:]:
            output[column] = output[column].map(lambda value: f"{value:.1f}")
    elif table_id == "Table4":
        output = data.copy()
        output.columns = ["Metric", "Value"]
        output["Metric"] = output["Metric"].replace(
            {
                "Subsample ARI (k=3) median": "k=3 subsample ARI, median",
                "Subsample ARI (k=3) p10": "k=3 subsample ARI, 10th percentile",
                "k=3 Stable Core": "k=3 stable core",
                "k=3 Moderate": "k=3 moderate",
                "k=3 Boundary": "k=3 boundary",
                "Ward k=3 ARI": "Ward versus k-means, k=3 ARI",
                "GMM k=3 ARI": "Gaussian mixture versus k-means, k=3 ARI",
            }
        )
    elif table_id == "Table5":
        output = pd.DataFrame(
            [
                ("Valid n", "787"),
                ("χ² (df=16)", "437.1"),
                ("Raw Cramér’s V", "0.5270"),
                ("Bias-corrected Cramér’s V", "0.5172"),
                ("95% bootstrap CI", "0.4898–0.5758"),
                ("Permutation p", "<0.0001"),
                ("Sparse cells", "4/27"),
            ],
            columns=["Statistic", "Value"],
        )
        transformation_note = (
            "The accepted source rows were retained and display precision was "
            "synchronized to the resolved Figure 6 registry; χ² was added as the "
            "registered reporting statistic without recomputation."
        )
    elif table_id == "TableS1":
        output = data.copy()
        output.columns = [
            "k=4 cluster",
            "n",
            "Median MLCAPE (J kg−1)",
            "Median 0–6-km shear (m s−1)",
            "Median 0–1-km SRH (m² s−2)",
        ]
        output["k=4 cluster"] = output["k=4 cluster"].map(lambda value: f"K4-C{int(value)}")
        for column in output.columns[2:]:
            output[column] = output[column].map(lambda value: f"{value:.3f}")
    elif table_id == "TableS2":
        output = data.copy()
        output.columns = ["Analysis", "Value"]
        output["Analysis"] = output["Analysis"].replace(
            {
                "k=3 stable core": "k=3 stable core",
                "k=4 stable core": "k=4 stable core",
                "Ward k=3 ARI": "Ward versus k-means, k=3 ARI",
                "Ward k=4 ARI": "Ward versus k-means, k=4 ARI",
                "GMM k=3 ARI": "Gaussian mixture versus k-means, k=3 ARI",
                "GMM k=4 ARI": "Gaussian mixture versus k-means, k=4 ARI",
            }
        )
    elif table_id == "TableS3":
        output = data.copy()
        output.columns = ["Metric", "Value"]
        output["Metric"] = output["Metric"].replace(
            {
                "STP_mod AUC": "STP_mod AUC",
                "SRH1 rated/unrated Cliff delta": "SRH1 rated/unrated Cliff’s δ",
            }
        )
        output.loc[output["Metric"] == "Status", "Value"] = "Exploratory; research-only"
    else:
        raise KeyError(table_id)
    return output.astype(str), transformation_note


def _markdown_table(data: pd.DataFrame) -> str:
    def escape(value: object) -> str:
        return str(value).replace("|", "\\|").replace("\n", "<br>")

    header = "| " + " | ".join(escape(column) for column in data.columns) + " |"
    rule = "| " + " | ".join("---" for _ in data.columns) + " |"
    rows = [
        "| " + " | ".join(escape(value) for value in row) + " |"
        for row in data.itertuples(index=False, name=None)
    ]
    return "\n".join([header, rule, *rows])


def _write_xlsx(path: Path, data: pd.DataFrame, title: str, note: str) -> None:
    fd, temporary_name = tempfile.mkstemp(
        dir=path.parent, prefix=f".{path.stem}.", suffix=".xlsx"
    )
    os.close(fd)
    temporary = Path(temporary_name)
    try:
        with pd.ExcelWriter(temporary, engine="openpyxl") as writer:
            data.to_excel(writer, sheet_name="Table", index=False, startrow=2)
            sheet = writer.book["Table"]
            sheet["A1"] = title
            sheet["A1"].font = Font(name="Arial", bold=True, size=12)
            sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(data.columns))
            header_fill = PatternFill("solid", fgColor="DCE6EF")
            thin = Side(style="thin", color="B8B8B8")
            for cell in sheet[3]:
                cell.font = Font(name="Arial", bold=True, size=10)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = Border(bottom=thin)
            for row in sheet.iter_rows(min_row=4, max_row=3 + len(data), max_col=len(data.columns)):
                for cell in row:
                    cell.font = Font(name="Arial", size=9)
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            note_row = 4 + len(data)
            sheet.cell(note_row, 1, f"Note: {note}")
            sheet.cell(note_row, 1).font = Font(name="Arial", italic=True, size=9)
            sheet.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=len(data.columns))
            sheet.freeze_panes = "A4"
            for index, column in enumerate(data.columns, start=1):
                values = [str(column), *data[column].astype(str).tolist()]
                width = min(42, max(11, max(len(value) for value in values) * 1.05))
                sheet.column_dimensions[chr(64 + index)].width = width
            sheet.row_dimensions[1].height = 22
            sheet.row_dimensions[3].height = 32
        os.replace(temporary, path)
    finally:
        if temporary.exists():
            temporary.unlink()


def _wrapped_preview_data(data: pd.DataFrame) -> pd.DataFrame:
    result = data.copy()
    n_columns = len(result.columns)
    if n_columns <= 2:
        widths = [38, 24]
    elif n_columns <= 4:
        widths = [24] * n_columns
    else:
        widths = [15] * n_columns
    for column, width in zip(result.columns, widths):
        result[column] = result[column].map(
            lambda value: "\n".join(
                textwrap.wrap(str(value), width=width, break_long_words=False)
            )
        )
    return result


def _save_preview(path: Path, data: pd.DataFrame, title: str, note: str) -> dict:
    display = _wrapped_preview_data(data)
    row_lines = [
        max(1, max(str(value).count("\n") + 1 for value in row))
        for row in display.itertuples(index=False, name=None)
    ]
    height_mm = max(70, 34 + 10 * (1 + sum(row_lines)) + 18)
    n_columns = len(display.columns)
    lengths = []
    for column in display.columns:
        longest = max([len(str(column)), *[len(str(v).split("\n")[0]) for v in display[column]]])
        lengths.append(max(7, min(30, longest)))
    total = sum(lengths)
    col_widths = [length / total for length in lengths]

    with publication_style():
        fig, ax = plt.subplots(figsize=(mm_to_in(178), mm_to_in(height_mm)))
        fig.subplots_adjust(left=0.025, right=0.975, bottom=0.12, top=0.86)
        ax.set_axis_off()
        table = ax.table(
            cellText=display.values,
            colLabels=list(display.columns),
            cellLoc="left",
            colLoc="center",
            colWidths=col_widths,
            bbox=[0, 0, 1, 1],
        )
        table.auto_set_font_size(False)
        table.set_fontsize(7.0 if n_columns <= 4 else 6.2)
        total_units = 1.25 + sum(row_lines)
        header_height = 1.25 / total_units
        for col_index in range(n_columns):
            cell = table[(0, col_index)]
            cell.set_height(header_height)
            cell.set_facecolor("#DCE6EF")
            cell.set_edgecolor("#7F8C8D")
            cell.set_linewidth(0.7)
            cell.get_text().set_fontproperties(font_properties(bold=True, size=7.0 if n_columns <= 4 else 6.1))
            cell.get_text().set_ha("center")
        for row_index, line_count in enumerate(row_lines, start=1):
            height = line_count / total_units
            for col_index in range(n_columns):
                cell = table[(row_index, col_index)]
                cell.set_height(height)
                cell.set_edgecolor("#B8B8B8")
                cell.set_linewidth(0.45)
                cell.set_facecolor("#FAFAFA" if row_index % 2 == 0 else "white")
                cell.get_text().set_fontproperties(font_properties(size=6.9 if n_columns <= 4 else 6.0))
                cell.get_text().set_va("center")
        fig.text(0.025, 0.95, title, ha="left", va="top", fontsize=9.0, fontweight="bold")
        fig.text(
            0.025,
            0.035,
            "Note: " + "\n".join(textwrap.wrap(note, width=126, break_long_words=False)),
            ha="left",
            va="bottom",
            fontsize=6.6,
            color="#4F4F4F",
            wrap=True,
        )
        fd, temporary_name = tempfile.mkstemp(
            dir=path.parent, prefix=f".{path.stem}.", suffix=".png"
        )
        os.close(fd)
        temporary = Path(temporary_name)
        try:
            fig.savefig(temporary, format="png", dpi=250, facecolor="white", transparent=False)
            with Image.open(temporary) as image:
                image.convert("RGB").save(path, format="PNG", dpi=(250, 250))
        finally:
            if temporary.exists():
                temporary.unlink()
        plt.close(fig)
    with Image.open(path) as image:
        return {
            "path": relative_to_final(path),
            "sha256": sha256_file(path),
            "mode": image.mode,
            "dpi": [round(float(v), 2) for v in image.info.get("dpi", (0, 0))],
            "pixel_width": image.width,
            "pixel_height": image.height,
        }


def build_all_tables() -> list[dict]:
    ensure_directories()
    records = []
    for definition in TABLE_DEFINITIONS:
        data, transformation_note = _load_and_format(definition)
        root = MAIN_TABLE_DIR if definition["scope"] == "main" else SUPP_TABLE_DIR
        csv_path = root / "csv" / f"{definition['stem']}.csv"
        xlsx_path = root / "xlsx" / f"{definition['stem']}.xlsx"
        md_path = root / "markdown" / f"{definition['stem']}.md"
        preview_path = root / "previews" / f"{definition['stem']}_preview.png"
        data.to_csv(csv_path, index=False, encoding="utf-8-sig")
        _write_xlsx(xlsx_path, data, definition["title"], definition["note"])
        write_text(
            md_path,
            f"# {definition['id']}. {definition['title']}\n\n"
            f"{_markdown_table(data)}\n\n"
            f"Note: {definition['note']}\n",
        )
        preview_meta = _save_preview(preview_path, data, definition["title"], definition["note"])
        records.append(
            {
                "id": definition["id"],
                "scope": definition["scope"],
                "title": definition["title"],
                "note": definition["note"],
                "source": definition["source"].resolve().relative_to(PROJECT_ROOT).as_posix(),
                "source_sha256": sha256_file(definition["source"]),
                "transformation_note": transformation_note,
                "files": {
                    "csv": relative_to_final(csv_path),
                    "xlsx": relative_to_final(xlsx_path),
                    "markdown": relative_to_final(md_path),
                    "preview": relative_to_final(preview_path),
                },
                "sha256": {
                    "csv": sha256_file(csv_path),
                    "xlsx": sha256_file(xlsx_path),
                    "markdown": sha256_file(md_path),
                    "preview": sha256_file(preview_path),
                },
                "preview_metadata": preview_meta,
            }
        )
    write_json(QC_DIR / "table_build_metadata.json", records)
    return records


if __name__ == "__main__":
    build_all_tables()
